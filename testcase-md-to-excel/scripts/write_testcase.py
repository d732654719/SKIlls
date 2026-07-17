# -*- coding: utf-8 -*-
"""
测试用例Excel写入脚本

功能：
- 将测试点扩展为标准测试用例格式并写入Excel
- 支持序号自动生成（1,2,3...）
- 自动设置单元格格式（框线、换行、对齐）
"""

import sys
import openpyxl
from openpyxl.styles import Border, Side, Alignment


def write_testcases(excel_path, testcases, sheet_name='testcase'):
    """
    写入测试用例到Excel
    
    参数:
        excel_path: Excel文件路径
        testcases: 测试用例列表，每个用例为 [测试点, 测试内容, 测试条件, 测试步骤, 预期结果]
        sheet_name: 工作表名称，默认'testcase'
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name]
    
    # 样式定义
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # 清空现有数据（保留表头）
    max_row = ws.max_row
    if max_row > 1:
        ws.delete_rows(2, max_row - 1)
    
    # 写入数据
    for idx, tc in enumerate(testcases, 1):
        row_num = idx + 1
        
        # A列：序号
        ws.cell(row=row_num, column=1, value=idx)
        ws.cell(row=row_num, column=1).border = thin_border
        ws.cell(row=row_num, column=1).alignment = center_align
        
        # B-F列：测试点、测试内容、测试条件、测试步骤、预期结果
        for col, value in enumerate(tc, 2):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            cell.alignment = left_align
    
    # 表头也加框线
    for col in range(1, 7):
        ws.cell(row=1, column=col).border = thin_border
        ws.cell(row=1, column=col).alignment = center_align
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 30
    
    wb.save(excel_path)
    print(f'写入完成！共{len(testcases)}条用例，序号1-{len(testcases)}')
    return len(testcases)


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('用法: python write_testcase.py <excel_path>')
        print('或: python write_testcase.py <excel_path> "<testcases_json>"')
        sys.exit(1)
    
    excel_path = sys.argv[1]
    
    if len(sys.argv) >= 3:
        import json
        testcases = json.loads(sys.argv[2])
        write_testcases(excel_path, testcases)
    else:
        print('请提供测试用例数据（JSON格式）')
