# -*- coding: utf-8 -*-
"""
MD测试用例表格 → Excel 转换脚本

功能：
- 读取MD文件中的Markdown表格
- 解析表头和数据行
- 写入Excel并自动设置格式（框线、换行、对齐、字体）
"""

import sys
import os
import re
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font


def _is_separator_row(row_str):
    """判断是否是MD表格的分隔行，如 |---|---|"""
    cleaned = row_str.strip('|').strip()
    return bool(re.match(r'^[\s\-:]+(\|[\s\-:]+)*$', cleaned))


def _parse_row_cells(row_str):
    """解析表格行，返回单元格列表"""
    content = row_str.strip('|')
    temp_marker = "__TEMP_PIPE__"
    content = re.sub(r'\\\|', temp_marker, content)
    cells = content.split('|')
    return [cell.strip().replace(temp_marker, '|') for cell in cells]


def _extract_table_blocks(lines):
    """从所有行中提取表格块，每个块是连续的表格行"""
    blocks = []
    current_block = []
    for line in lines:
        stripped = line.strip()
        if stripped.startswith('|') and stripped.endswith('|'):
            current_block.append(stripped)
        else:
            if current_block:
                blocks.append(current_block)
                current_block = []
    if current_block:
        blocks.append(current_block)
    return blocks


def parse_md_table(md_path):
    """解析MD文件中的Markdown表格，合并所有同结构表格，返回 (headers, rows)"""
    with open(md_path, 'r', encoding='utf-8') as f:
        content = f.read()

    lines = content.strip().split('\n')
    blocks = _extract_table_blocks(lines)

    valid_tables = []
    for block in blocks:
        if len(block) < 3:
            continue
        if not _is_separator_row(block[1]):
            continue
        headers = _parse_row_cells(block[0])
        data_rows = [_parse_row_cells(row) for row in block[2:]]
        valid_tables.append((headers, data_rows))

    if not valid_tables:
        raise ValueError("未找到有效的MD表格（至少需要表头行、分隔行和数据行）")

    # 找出所有包含 testNo 列的表，合并数据行
    testcase_tables = []
    for headers, rows in valid_tables:
        if any('testno' in h.lower() or 'testNo' in h for h in headers):
            testcase_tables.append((headers, rows))

    if testcase_tables:
        # 使用第一个表的表头，合并所有表的数据行
        headers = testcase_tables[0][0]
        all_rows = []
        for _, rows in testcase_tables:
            all_rows.extend(rows)
    else:
        # 没有 testNo 列则选列数最多的表
        headers, all_rows = max(valid_tables, key=lambda t: len(t[0]))

    return headers, all_rows


def write_to_excel(headers, rows, xlsx_path, sheet_name='testcase'):
    """将数据写入Excel并设置格式"""
    wb = openpyxl.Workbook()
    
    # 第一个sheet：测试用例
    ws = wb.active
    ws.title = sheet_name

    # 样式定义
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    font_yahei = Font(name='微软雅黑', size=11)

    # 写入表头（第1行）
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.border = thin_border
        cell.alignment = center_align
        cell.font = font_yahei

    # 写入数据行（从第2行开始）
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            # 将 <br> 转换为换行符
            if isinstance(value, str):
                value = value.replace('<br>', '\n')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.font = font_yahei
            # 第一列（序号）居中，其余左对齐
            if col_idx == 1:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # 设置列宽
    col_widths = {
        'A': 8,   # 序号
        'B': 30,  # 测试内容
        'C': 25,  # 测试点
        'D': 30,  # 测试条件
        'E': 40,  # 测试步骤
        'F': 30,  # 测试结果
        'G': 21.13,  # 测试截图
        'H': 21.13,  # 测试截图2
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # 第二个sheet：测试结果记录（完全为空）
    ws_result = wb.create_sheet(title='caseResult')

    wb.save(xlsx_path)
    print(f'写入完成！表头: {len(headers)}列，数据: {len(rows)}行 → {xlsx_path}')
    print(f'已创建第二个空sheet: caseResult')
    return len(rows)


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print('用法: python md_to_excel.py <md文件路径> <输出xlsx路径>')
        print('示例: python md_to_excel.py "E:/项目/XX项目/XX-用例.md" "E:/项目/XX项目/XX-用例.xlsx"')
        sys.exit(1)

    md_path = sys.argv[1]
    xlsx_path = sys.argv[2]

    if not os.path.exists(md_path):
        print(f'错误: 文件不存在 - {md_path}')
        sys.exit(1)

    headers, rows = parse_md_table(md_path)
    count = write_to_excel(headers, rows, xlsx_path)
    print(f'转换完成！共 {count} 条用例。')
